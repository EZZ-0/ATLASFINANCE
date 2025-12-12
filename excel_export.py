"""
Excel Export Module with Formula Preservation
Creates professional Excel workbooks with formatting, formulas, and multiple sheets
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from datetime import datetime


def _convert_numpy_types(value):
    """
    Convert numpy types to Python native types for Excel compatibility.
    openpyxl cannot serialize numpy types directly.
    """
    if value is None:
        return None
    
    # Handle numpy scalar types
    if isinstance(value, (np.integer, np.int64, np.int32)):
        return int(value)
    elif isinstance(value, (np.floating, np.float64, np.float32)):
        return float(value)
    elif isinstance(value, np.bool_):
        return bool(value)
    elif isinstance(value, np.ndarray):
        return value.tolist()
    elif isinstance(value, dict):
        # Recursively convert dict values
        return {k: _convert_numpy_types(v) for k, v in value.items()}
    elif isinstance(value, (list, tuple)):
        return [_convert_numpy_types(v) for v in value]
    
    return value


class FinancialExcelExporter:
    """
    Professional Excel export with formatting and formulas
    """
    
    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
    
    def _format_header(self, ws, row=1):
        """Apply header formatting"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _format_currency_column(self, ws, col_letter, start_row=2):
        """Format column as currency"""
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    
    def _format_percentage_column(self, ws, col_letter, start_row=2):
        """Format column as percentage"""
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'
    
    def add_financial_statement(self, df, sheet_name="Financials"):
        """
        Add financial statement with proper orientation and formatting
        Metrics as ROWS, Dates as COLUMNS
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Ensure correct orientation: metrics as rows, dates as columns
        if len(df.index) > 0:
            first_idx = df.index[0]
            # If index is date, transpose so metrics become rows
            if isinstance(first_idx, pd.Timestamp):
                df = df.T
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Convert numpy types to Python native types for Excel compatibility
                value = _convert_numpy_types(value)
                
                # Try to convert strings back to numbers for Excel
                if r_idx > 2 and value is not None:  # Skip headers
                    try:
                        if isinstance(value, str):
                            # Remove common formatting characters
                            cleaned = value.replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
                            # Remove unit suffixes (B, M, K, T)
                            if cleaned and cleaned[-1] in ['B', 'M', 'K', 'T']:
                                multiplier = {'T': 1e12, 'B': 1e9, 'M': 1e6, 'K': 1e3}
                                value = float(cleaned[:-1]) * multiplier.get(cleaned[-1], 1)
                            # Try direct conversion
                            elif cleaned.replace('.','',1).replace('-','',1).isdigit():
                                value = float(cleaned)
                    except (ValueError, AttributeError):
                        pass
                
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # ONLY format if value is actually a number
                if isinstance(value, (int, float)) and not isinstance(value, bool) and r_idx > 2:
                    if abs(value) >= 1:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00%'
        
        # Format header
        self._format_header(ws, row=1)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        return ws
    
    def add_dcf_model(self, projections, valuation_summary, sheet_name="DCF Model"):
        """
        Add DCF model with formulas preserved
        Users can modify assumptions and formulas will update
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Section 1: Valuation Summary
        ws['A1'] = "DCF Valuation Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        for key, value in valuation_summary.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Section 2: Cash Flow Projections
        row += 2
        ws[f'A{row}'] = "5-Year Cash Flow Projections"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 2
        proj_start_row = row
        
        # Write projections
        for r_idx, data_row in enumerate(dataframe_to_rows(projections, index=False, header=True), row):
            for c_idx, value in enumerate(data_row, 1):
                # Convert numpy types for Excel compatibility
                value = _convert_numpy_types(value)
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format large numbers
                if isinstance(value, (int, float)) and value > 1000:
                    cell.number_format = '#,##0.00'
        
        # Format headers
        self._format_header(ws, row=proj_start_row)
        
        # Add formulas for growth rates
        growth_row = proj_start_row + len(projections) + 2
        ws[f'A{growth_row}'] = "YoY Growth %"
        ws[f'A{growth_row}'].font = Font(bold=True)
        
        # Revenue growth formula
        for col_idx in range(2, 7):  # Years 1-5
            col_letter = chr(65 + col_idx)  # B, C, D, E, F
            prev_col = chr(65 + col_idx - 1)
            formula = f"=({col_letter}{proj_start_row+1}-{prev_col}{proj_start_row+1})/{prev_col}{proj_start_row+1}"
            cell = ws[f'{col_letter}{growth_row}']
            cell.value = formula
            cell.number_format = '0.00%'
        
        return ws
    
    def add_comparison_sheet(self, comparison_df, sheet_name="Company Comparison"):
        """Add multi-company comparison"""
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Convert numpy types for Excel compatibility
                value = _convert_numpy_types(value)
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format header
        self._format_header(ws, row=1)
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        return ws
    
    def add_cover_sheet(self, ticker, company_name, extraction_time):
        """Add professional cover sheet"""
        ws = self.wb.create_sheet(title="Cover", index=0)
        
        # Title
        ws['A1'] = "USA EARNINGS ENGINE"
        ws['A1'].font = Font(bold=True, size=20, color="366092")
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Financial Analysis Report"
        ws['A3'].font = Font(size=14)
        ws.merge_cells('A3:D3')
        
        # Company info
        ws['A6'] = "Company:"
        ws['B6'] = company_name
        ws['B6'].font = Font(bold=True)
        
        ws['A7'] = "Ticker:"
        ws['B7'] = ticker
        ws['B7'].font = Font(bold=True)
        
        ws['A8'] = "Report Date:"
        ws['B8'] = datetime.now().strftime("%B %d, %Y")
        
        ws['A9'] = "Extraction Time:"
        ws['B9'] = f"{extraction_time:.2f}s"
        
        # Disclaimer
        ws['A12'] = "DISCLAIMER:"
        ws['A12'].font = Font(bold=True, size=11)
        ws['A13'] = "This report is for informational purposes only. Not investment advice."
        ws['A13'].font = Font(italic=True, size=9)
        
        return ws
    
    def save(self, filename):
        """Save workbook to file"""
        self.wb.save(filename)
        print(f"[OK] Excel workbook saved: {filename}")


def export_financials_to_excel(financials: dict, output_file: str):
    """
    Main export function
    
    Args:
        financials: Dictionary from st.session_state.financials
        output_file: Path to save .xlsx file
    
    Returns:
        Path to saved file
    """
    exporter = FinancialExcelExporter()
    
    # Cover sheet
    extraction_time = financials.get("extraction_time", 0)
    # Handle if extraction_time is a string like "10.25s"
    if isinstance(extraction_time, str):
        try:
            extraction_time = float(extraction_time.replace('s', '').strip())
        except ValueError:
            extraction_time = 0
    
    exporter.add_cover_sheet(
        ticker=financials.get("ticker", "N/A"),
        company_name=financials.get("company_name", "N/A"),
        extraction_time=extraction_time
    )
    
    # Financial statements
    income = financials.get("income_statement", pd.DataFrame())
    if not income.empty:
        exporter.add_financial_statement(income, "Income Statement")
    
    balance = financials.get("balance_sheet", pd.DataFrame())
    if not balance.empty:
        exporter.add_financial_statement(balance, "Balance Sheet")
    
    cashflow = financials.get("cash_flow", pd.DataFrame())
    if not cashflow.empty:
        exporter.add_financial_statement(cashflow, "Cash Flow")
    
    # Ratios
    ratios = financials.get("ratios", pd.DataFrame())
    if not ratios.empty:
        # Filter out rows containing dict values (like _components) that can't be serialized to Excel
        ratios_clean = ratios.copy()
        rows_to_drop = [idx for idx in ratios_clean.index 
                        if isinstance(ratios_clean.loc[idx].iloc[0] if hasattr(ratios_clean.loc[idx], 'iloc') 
                                      else ratios_clean.loc[idx], dict)]
        if rows_to_drop:
            ratios_clean = ratios_clean.drop(rows_to_drop)
        exporter.add_financial_statement(ratios_clean, "Financial Ratios")
    
    # Save
    exporter.save(output_file)
    
    return output_file


# Test function
if __name__ == "__main__":
    print("Testing Excel export...")
    
    # Create sample data
    sample_income = pd.DataFrame({
        'Total Revenue': [145000000000, 135000000000, 126000000000],
        'Net Income': [11200000000, 9370000000, 9700000000],
        'EBIT': [133000000000, 123000000000, 114000000000]
    }, index=['2025-01-31', '2024-01-31', '2023-01-31'])
    
    sample_financials = {
        "ticker": "TEST",
        "company_name": "Test Company Inc.",
        "extraction_time": 0.5,
        "income_statement": sample_income
    }
    
    export_financials_to_excel(sample_financials, "test_export.xlsx")
    print("[OK] Test export completed!")

